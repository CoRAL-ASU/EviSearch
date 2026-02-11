#!/usr/bin/env python3
"""
Aggregate accuracy scores and results for all trials in new_pipeline_outputs/results.
Produces:
  - Excel report with Summary sheet + Per-Column sheet (wrong values highlighted in red)
  - HTML + PNG visualizations (overall scores per trial, heatmap of column correctness)

Usage:
  python aggregate_accuracy_report.py                    # use default: new_pipeline_outputs/results
  python aggregate_accuracy_report.py "new_pipeline_outputs copy/results"   # use alternate folder
"""
import json
import sys
from pathlib import Path

# Project root
REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RESULTS_DIR = REPO_ROOT / "new_pipeline_outputs" / "results"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "new_pipeline_outputs" / "accuracy_report"

# Set by main() from CLI; load_all_trial_evaluations uses RESULTS_DIR
RESULTS_DIR = DEFAULT_RESULTS_DIR
OUTPUT_DIR = DEFAULT_OUTPUT_DIR


def find_evaluation_dir(trial_dir: Path) -> Path | None:
    """Return path to evaluation dir for this trial, or None."""
    # Prefer: trial/evaluation > trial/latest/evaluation > first trial/run_*/evaluation
    direct = trial_dir / "evaluation"
    if (direct / "evaluation_results.json").exists():
        return direct
    latest = trial_dir / "latest" / "evaluation"
    if (latest / "evaluation_results.json").exists():
        return latest
    for run in sorted(trial_dir.iterdir()):
        if run.is_dir() and run.name.startswith("run_"):
            ev = run / "evaluation"
            if (ev / "evaluation_results.json").exists():
                return ev
    return None


def load_all_trial_evaluations() -> dict:
    """Discover all trials with evaluation and load their results. Returns {trial_name: {summary, columns}}. """
    data = {}
    if not RESULTS_DIR.exists():
        print(f"Results dir not found: {RESULTS_DIR}")
        return data

    for trial_dir in sorted(RESULTS_DIR.iterdir()):
        if not trial_dir.is_dir():
            continue
        trial_name = trial_dir.name
        eval_dir = find_evaluation_dir(trial_dir)
        if not eval_dir:
            continue

        results_path = eval_dir / "evaluation_results.json"
        summary_path = eval_dir / "summary_metrics.json"
        if not results_path.exists():
            continue

        with open(results_path, "r", encoding="utf-8") as f:
            results = json.load(f)
        summary = None
        if summary_path.exists():
            with open(summary_path, "r", encoding="utf-8") as f:
                summary = json.load(f)

        data[trial_name] = {
            "summary": summary or {},
            "columns": results.get("columns", {}),
            "document_name": results.get("document_name", trial_name),
        }
        print(f"  Loaded: {trial_name}")

    return data


def build_summary_df(data: dict):
    """DataFrame: one row per trial with overall and by-category metrics.
    Supports both formats:
    - New: summary.summary = {Document, Total columns, Correct, Partial, Wrong, avg_overall};
           summary.by_category = list of {Category, Total, Correct, Partial, Wrong}
    - Old: summary.overall = {avg_correctness, avg_completeness, avg_overall, total_columns};
           summary.by_category = dict of category -> {avg_overall, column_count}
    """
    import pandas as pd

    rows = []
    for trial, payload in data.items():
        s = payload.get("summary") or {}
        row = {"Trial": trial}

        # New format (evaluator_v2 with verdict)
        if "summary" in s and isinstance(s["summary"], dict):
            sm = s["summary"]
            row["avg_overall"] = sm.get("avg_overall")
            row["total_columns"] = sm.get("Total columns")
            row["Correct"] = sm.get("Correct")
            row["Partial"] = sm.get("Partial")
            row["Wrong"] = sm.get("Wrong")
            row["avg_correctness"] = sm.get("avg_overall")  # reuse for display
            row["avg_completeness"] = sm.get("avg_overall")
            by_cat = s.get("by_category") or []
            if isinstance(by_cat, list):
                by_cat_dict = {r["Category"]: r for r in by_cat if isinstance(r, dict)}
                for cat in ["exact_match", "numeric_tolerance", "structured_text"]:
                    c = by_cat_dict.get(cat) or {}
                    tot = c.get("Total", 0)
                    row[f"{cat}_count"] = tot
                    row[f"{cat}_overall"] = (
                        (c.get("Correct", 0) * 1.0 + c.get("Partial", 0) * 0.5) / tot
                        if tot else None
                    )
            else:
                for cat in ["exact_match", "numeric_tolerance", "structured_text"]:
                    row[f"{cat}_overall"] = None
                    row[f"{cat}_count"] = None
        else:
            # Old format
            overall = s.get("overall") or {}
            by_cat = s.get("by_category") or {}
            row["avg_correctness"] = overall.get("avg_correctness")
            row["avg_completeness"] = overall.get("avg_completeness")
            row["avg_overall"] = overall.get("avg_overall")
            row["total_columns"] = overall.get("total_columns")
            row["Correct"] = None
            row["Partial"] = None
            row["Wrong"] = None
            for cat in ["exact_match", "numeric_tolerance", "structured_text"]:
                c = by_cat.get(cat) or {}
                row[f"{cat}_overall"] = c.get("avg_overall")
                row[f"{cat}_count"] = c.get("column_count")
        rows.append(row)
    return pd.DataFrame(rows)


def build_per_column_df(data: dict):
    """DataFrame: rows = column names (union), columns = trial names, values = overall score (0-1). Wrong = < 1.0."""
    import pandas as pd

    all_columns = set()
    for payload in data.values():
        all_columns.update((payload.get("columns") or {}).keys())
    all_columns = sorted(all_columns)

    rows = []
    for col in all_columns:
        row = {"Column": col}
        for trial in data:
            cols_data = (data[trial].get("columns") or {}).get(col)
            if cols_data is not None:
                row[trial] = cols_data.get("overall", None)
            else:
                row[trial] = None  # column not present in this trial
        rows.append(row)

    return pd.DataFrame(rows)


def build_wrong_only_df(data: dict):
    """DataFrame: one row per (trial, column) where overall < 1.0, with GT and Pred."""
    import pandas as pd

    rows = []
    for trial, payload in data.items():
        cols = payload.get("columns") or {}
        for col_name, col_data in cols.items():
            overall = col_data.get("overall")
            if overall is not None and overall < 1.0:
                rows.append({
                    "Trial": trial,
                    "Column": col_name,
                    "verdict": col_data.get("verdict"),  # Correct/Partial/Wrong when present
                    "overall": overall,
                    "correctness": col_data.get("correctness"),
                    "completeness": col_data.get("completeness"),
                    "category": col_data.get("category"),
                    "ground_truth": str(col_data.get("ground_truth", ""))[:200],
                    "predicted": str(col_data.get("predicted", ""))[:200],
                    "reason": (col_data.get("reason") or "")[:300],
                })
    return pd.DataFrame(rows)


def write_excel(data: dict, summary_df, per_column_df, wrong_df, path: Path):
    """Write Excel with Summary, Per-Column (red for wrong), and Wrong-Only sheets."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("openpyxl not installed. Saving as CSV instead. Install with: pip install openpyxl")
        summary_df.to_csv(path.with_suffix(".summary.csv"), index=False)
        per_column_df.to_csv(path.with_suffix(".per_column.csv"), index=False)
        wrong_df.to_csv(path.with_suffix(".wrong_only.csv"), index=False)
        return

    red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    wb = openpyxl.Workbook()
    # Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws_sum.append(r)
    for cell in ws_sum[1]:
        cell.font = Font(bold=True)

    # Per-Column: highlight wrong (< 1.0) in red
    ws_col = wb.create_sheet("Per-Column Scores")
    for r in dataframe_to_rows(per_column_df, index=False, header=True):
        ws_col.append(r)
    for cell in ws_col[1]:
        cell.font = Font(bold=True)
    # Data starts at row 2; column A = "Column", then one column per trial
    trial_cols = list(per_column_df.columns)
    trial_cols.remove("Column")
    for row_idx, row in enumerate(ws_col.iter_rows(min_row=2, max_row=ws_col.max_row), start=2):
        for col_idx, cell in enumerate(row, start=1):
            if col_idx == 1:
                continue
            if cell.value is not None:
                try:
                    v = float(cell.value)
                    if v < 1.0:
                        cell.fill = red_fill
                    elif v == 1.0:
                        cell.fill = green_fill
                except (TypeError, ValueError):
                    pass

    # Wrong only
    ws_wrong = wb.create_sheet("Wrong Only")
    for r in dataframe_to_rows(wrong_df, index=False, header=True):
        ws_wrong.append(r)
    for cell in ws_wrong[1]:
        cell.font = Font(bold=True)

    wb.save(path)
    print(f"Excel saved: {path}")


def create_visualizations(data: dict, summary_df, per_column_df, out_dir: Path):
    """Create bar chart (overall per trial) and heatmap (column x trial, green/red)."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np
    except ImportError:
        print("matplotlib not installed. Skipping PNG visualizations. Install with: pip install matplotlib")
        return

    out_dir.mkdir(parents=True, exist_ok=True)

    # 1) Bar chart: avg_overall per trial
    fig, ax = plt.subplots(figsize=(10, 5))
    trials = summary_df["Trial"].tolist()
    scores = summary_df["avg_overall"].tolist()
    colors = ["#2ecc71" if s >= 0.7 else "#e74c3c" if s < 0.5 else "#f39c12" for s in scores]
    x = range(len(trials))
    bars = ax.bar(x, scores, color=colors, edgecolor="black", linewidth=0.5)
    ax.set_xticks(x)
    ax.set_xticklabels([t.replace("'", "'\n") for t in trials], rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Average overall score")
    ax.set_title("Accuracy (avg overall) per trial")
    ax.set_ylim(0, 1.05)
    ax.axhline(y=0.7, color="gray", linestyle="--", alpha=0.7)
    plt.tight_layout()
    fig.savefig(out_dir / "accuracy_per_trial.png", dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Saved: {out_dir / 'accuracy_per_trial.png'}")

    # 2) Heatmap: columns x trials (overall score); red = wrong, green = correct
    # Use a subset of columns if too many (e.g. max 80 columns for readability)
    df = per_column_df.set_index("Column")
    trial_names = [c for c in df.columns if c in data]
    if not trial_names:
        return
    plot_df = df[trial_names].fillna(-1)  # missing = -1 (will show as gray)
    if len(plot_df) > 80:
        # Keep columns that have at least one wrong in any trial
        wrong_mask = (plot_df < 1.0) & (plot_df >= 0)
        cols_with_wrong = wrong_mask.any(axis=1)
        other = plot_df.index.difference(plot_df.index[cols_with_wrong])[: 80 - cols_with_wrong.sum()]
        plot_df = plot_df.loc[list(plot_df.index[cols_with_wrong]) + list(other)]
    if plot_df.empty:
        return

    fig, ax = plt.subplots(figsize=(max(8, len(trial_names) * 1.2), max(6, len(plot_df) * 0.15)))
    # -1 = missing, 0-1 = score
    im = ax.imshow(plot_df.values, aspect="auto", cmap="RdYlGn", vmin=-0.05, vmax=1.05)
    ax.set_xticks(range(len(trial_names)))
    ax.set_xticklabels([t.replace("'", "'\n") for t in trial_names], rotation=45, ha="right", fontsize=7)
    ax.set_yticks(range(len(plot_df)))
    ax.set_yticklabels(list(plot_df.index), fontsize=6)
    ax.set_title("Per-column overall score (green=correct, red=wrong)")
    plt.colorbar(im, ax=ax, label="Score")
    plt.tight_layout()
    fig.savefig(out_dir / "heatmap_per_column_trial.png", dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Saved: {out_dir / 'heatmap_per_column_trial.png'}")


def create_html_report(data: dict, summary_df, per_column_df, wrong_df, out_dir: Path):
    """Single HTML page: summary table + per-trial breakdown + wrong values highlighted."""
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "accuracy_report.html"

    trials = list(data.keys())
    # Build per-column matrix for HTML table (column name, then one cell per trial with score and class)
    all_cols_set = set()
    for payload in data.values():
        all_cols_set.update((payload.get("columns") or {}).keys())
    all_cols = sorted(all_cols_set)

    def score_cell(trial: str, col: str) -> tuple[str, str]:
        cols = (data.get(trial) or {}).get("columns") or {}
        c = cols.get(col)
        if c is None:
            return "—", "missing"
        s = c.get("overall")
        if s is None:
            return "—", "missing"
        if s >= 1.0:
            return f"{s:.2f}", "correct"
        return f"{s:.2f}", "wrong"

    rows_html = []
    for col in all_cols:
        cells = [f'<td class="colname" title="{col[:100]}">{col[:40]}...</td>' if len(col) > 40 else f'<td class="colname">{col}</td>']
        for trial in trials:
            val, cls = score_cell(trial, col)
            cells.append(f'<td class="{cls}">{val}</td>')
        rows_html.append("<tr>" + "".join(cells) + "</tr>")

    summary_rows = []
    has_verdict = "Correct" in summary_df.columns and summary_df["Correct"].notna().any()
    for _, row in summary_df.iterrows():
        t = row["Trial"]
        avg = row.get("avg_overall")
        avg_str = f"{avg:.3f}" if avg is not None and not (isinstance(avg, float) and (avg != avg)) else "—"
        acc = row.get("avg_correctness")
        acc_str = f"{acc:.3f}" if acc is not None and not (isinstance(acc, float) and (acc != acc)) else "—"
        comp = row.get("avg_completeness")
        comp_str = f"{comp:.3f}" if comp is not None and not (isinstance(comp, float) and (comp != comp)) else "—"
        tot = row.get("total_columns")
        tot_str = str(int(tot)) if tot is not None else "—"
        if has_verdict and row.get("Correct") is not None:
            summary_rows.append(
                f"<tr><td>{t}</td><td>{avg_str}</td><td>{tot_str}</td>"
                f"<td>{int(row['Correct'])}</td><td>{int(row['Partial'])}</td><td>{int(row['Wrong'])}</td></tr>"
            )
        else:
            summary_rows.append(
                f"<tr><td>{t}</td><td>{avg_str}</td><td>{acc_str}</td><td>{comp_str}</td><td>{tot_str}</td></tr>"
            )

    if has_verdict:
        summary_header = "<thead><tr><th>Trial</th><th>Avg overall</th><th>Total columns</th><th>Correct</th><th>Partial</th><th>Wrong</th></tr></thead>"
    else:
        summary_header = "<thead><tr><th>Trial</th><th>Avg overall</th><th>Avg correctness</th><th>Avg completeness</th><th>Total columns</th></tr></thead>"

    trial_headers = "".join(f"<th>{t}</th>" for t in trials)
    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Accuracy Report – All Trials</title>
<style>
  body {{ font-family: sans-serif; margin: 20px; background: #f5f5f5; }}
  h1 {{ color: #333; }}
  h2 {{ margin-top: 24px; color: #444; }}
  table {{ border-collapse: collapse; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
  th, td {{ border: 1px solid #ddd; padding: 6px 10px; text-align: left; }}
  th {{ background: #333; color: white; }}
  td.correct {{ background: #C6EFCE; }}
  td.wrong {{ background: #FFCCCB; }}
  td.missing {{ background: #eee; color: #999; }}
  td.colname {{ max-width: 280px; font-size: 0.9em; }}
  .summary table {{ margin-bottom: 20px; }}
  .summary th {{ background: #2c3e50; }}
  .note {{ color: #666; font-size: 0.9em; margin-top: 8px; }}
</style>
</head>
<body>
<h1>Accuracy Report – All Trials</h1>
<p class="note">Generated from new_pipeline_outputs/results. Green = correct (1.0), Red = wrong (&lt; 1.0).</p>

<h2>Summary (avg overall per trial)</h2>
<div class="summary">
<table>
{summary_header}
<tbody>
{"".join(summary_rows)}
</tbody>
</table>
</div>

<h2>Per-column scores (all trials)</h2>
<p class="note">Scroll right to see all trials. Red = score &lt; 1.0.</p>
<div style="overflow-x: auto;">
<table>
<thead><tr><th>Column</th>{trial_headers}</tr></thead>
<tbody>
{"".join(rows_html)}
</tbody>
</table>
</div>
</body>
</html>
"""
    path.write_text(html, encoding="utf-8")
    print(f"HTML report saved: {path}")


def main():
    global RESULTS_DIR, OUTPUT_DIR
    import pandas as pd

    if len(sys.argv) > 1:
        arg = sys.argv[1].strip()
        results_path = Path(arg)
        if not results_path.is_absolute():
            results_path = REPO_ROOT / results_path
        if not results_path.is_dir():
            print(f"Results directory not found: {results_path}")
            sys.exit(1)
        RESULTS_DIR = results_path
        # Write report next to the results folder (e.g. new_pipeline_outputs copy/accuracy_report)
        OUTPUT_DIR = RESULTS_DIR.parent / "accuracy_report"
        print("Using results dir:", RESULTS_DIR)
        print("Output dir:", OUTPUT_DIR)
    else:
        RESULTS_DIR = DEFAULT_RESULTS_DIR
        OUTPUT_DIR = DEFAULT_OUTPUT_DIR

    print("Scanning for trials with evaluation results...")
    data = load_all_trial_evaluations()
    if not data:
        print("No evaluation data found.")
        sys.exit(1)

    print(f"Found {len(data)} trials: {list(data.keys())}")

    summary_df = build_summary_df(data)
    per_column_df = build_per_column_df(data)
    wrong_df = build_wrong_only_df(data)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    excel_path = OUTPUT_DIR / "accuracy_report.xlsx"
    write_excel(data, summary_df, per_column_df, wrong_df, excel_path)

    create_visualizations(data, summary_df, per_column_df, OUTPUT_DIR)
    create_html_report(data, summary_df, per_column_df, wrong_df, OUTPUT_DIR)

    print("\nDone. Outputs in:", OUTPUT_DIR)
    print("  - accuracy_report.xlsx (Summary, Per-Column with red highlight, Wrong Only)")
    print("  - accuracy_report.html (view in browser)")
    print("  - accuracy_per_trial.png, heatmap_per_column_trial.png")


if __name__ == "__main__":
    main()
